import mysql.connector
from dbconnection import DB_CONFIG
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
import json
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class BusinessReportExporter:
    def __init__(self):
        pass
    
    def export_business_report(self, time_filter="All Time", payment_info=None, order_type_info=None):
        
        
        """Download Complete Reports"""
        
        
        try:
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Ice'n Spice Business Reports"
            )
            
            if not file_path:
                return
                
            # Connect to database
            conn = mysql.connector.connect(**DB_CONFIG)
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Sheet 1: Sales analytics with graphs
            self._create_sales_analytics_sheet(conn, workbook, time_filter)
            
            # Sheet 2: Customer details
            self._create_customer_sheet(conn, workbook, time_filter)
            
            # Sheet 3: Orders details like in the order page
            self._create_orders_sheet(conn, workbook, time_filter, payment_info, order_type_info)
            
            # Sheet 4: Items list
            self._create_items_sheet(conn, workbook, time_filter)
            
            # Save workbook
            workbook.save(file_path)
            
            # Close database connection
            conn.close()
            
            # Show success message
            messagebox.showinfo("Export Complete", "Business reports exported successfully!")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export reports: {str(e)}")
    
    def _get_time_filter_clause(self, time_filter):
        if time_filter == "This Month":
            return "WHERE YEAR(o.CreatedAT) = YEAR(CURDATE()) AND MONTH(o.CreatedAT) = MONTH(CURDATE())"
        elif time_filter == "Last Month":
            return "WHERE YEAR(o.CreatedAT) = YEAR(DATE_SUB(CURDATE(), INTERVAL 1 MONTH)) AND MONTH(o.CreatedAT) = MONTH(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))"
        elif time_filter == "This Year":
            return "WHERE YEAR(o.CreatedAT) = YEAR(CURDATE())"
        else:  # "All Time"
            return ""

    def _create_sales_analytics_sheet(self, conn, workbook, time_filter):
        # Create sales data query with time filter
        time_clause = self._get_time_filter_clause(time_filter)
        sales_query = f"""
            SELECT 
                m.Category,
                MONTHNAME(o.CreatedAT) as Month,
                YEAR(o.CreatedAT) as Year,
                COUNT(*) as OrderCount,
                SUM(m.Price) as TotalRevenue,
                COUNT(DISTINCT o.UserName) as CustomerCount
            FROM Menu m
            JOIN `Order` o ON m.MenuID IN (
                SELECT JSON_EXTRACT(SUBSTRING_INDEX(SUBSTRING_INDEX(o.Item_list, '}}', 1), '{{', -1), '$.id')
            )
            {time_clause}
            GROUP BY m.Category, YEAR(o.CreatedAT), MONTHNAME(o.CreatedAT), MONTH(o.CreatedAT)
            ORDER BY YEAR(o.CreatedAT), MONTH(o.CreatedAT), TotalRevenue DESC
        """
        
        # Get data
        sales_df = pd.read_sql(sales_query, conn)
        
        # Create sales sheet
        sales_sheet = workbook.active
        sales_sheet.title = "Sales_Analytics"
        
        # Add title
        sales_sheet['A1'] = "Sales Analytics Report"
        sales_sheet['A1'].font = Font(size=16, bold=True)
        sales_sheet.merge_cells('A1:F1')
        sales_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        sales_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        sales_sheet['A2'].font = Font(size=10, italic=True)
        sales_sheet.merge_cells('A2:F2')
        sales_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Add header row at A4
        headers = ["Category", "Month", "Year", "Order Count", "Total Revenue", "Customer Count"]
        for col, header in enumerate(headers, start=1):
            cell = sales_sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F1D94B", end_color="F1D94B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        if not sales_df.empty:
            for r_idx, row in enumerate(sales_df.itertuples(), start=5):
                sales_sheet.cell(row=r_idx, column=1, value=row.Category)
                sales_sheet.cell(row=r_idx, column=2, value=row.Month)
                sales_sheet.cell(row=r_idx, column=3, value=row.Year)
                sales_sheet.cell(row=r_idx, column=4, value=row.OrderCount)
                sales_sheet.cell(row=r_idx, column=5, value=f"${row.TotalRevenue:.2f}")
                sales_sheet.cell(row=r_idx, column=6, value=row.CustomerCount)
        
        # Create graph if data exists
        if not sales_df.empty:
            # Create a pivot table for easier plotting
            pivot_df = sales_df.pivot_table(
                index='Month', 
                columns='Category', 
                values='TotalRevenue',
                aggfunc='sum'
            ).fillna(0)
            
            # Create graph
            fig, ax = plt.subplots(figsize=(10, 6))
            pivot_df.plot(kind='bar', ax=ax)
            ax.set_title('Sales Revenue by Category')
            ax.set_ylabel('Revenue ($)')
            ax.set_xlabel('Month')
            ax.legend(title='Category')
            
            # Save the figure to a temporary file
            temp_img_path = 'temp_sales_graph.png'
            plt.savefig(temp_img_path)
            plt.close()
            
            # Create a chart sheet
            chart_sheet = workbook.create_sheet(title="Sales_Charts")
            
            # Add chart title
            chart_sheet['A1'] = "Sales Revenue by Category"
            chart_sheet['A1'].font = Font(size=16, bold=True)
            chart_sheet.merge_cells('A1:J1')
            chart_sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Add image to Excel
            img = openpyxl.drawing.image.Image(temp_img_path)
            chart_sheet.add_image(img, 'A3')
            
            # Remove temporary file
            os.remove(temp_img_path)
        
        # Format sheet
        self._format_sheet(sales_sheet)
    
    def _create_customer_sheet(self, conn, workbook, time_filter):
        """Create customer details sheet 
            no passwords are there"""
        # Modify customer query to include time filter if needed
        time_clause = self._get_time_filter_clause(time_filter).replace('o.CreatedAT', 'created_at')
        customer_query = f"""
            SELECT 
                userID, 
                first_name, 
                last_name, 
                username, 
                email, 
                phone_number, 
                address, 
                role, 
                created_at
            FROM users
            {time_clause}
            ORDER BY created_at DESC
        """
        
        # Get data
        customer_df = pd.read_sql(customer_query, conn)
        
        # Create customer sheet
        customer_sheet = workbook.create_sheet(title="Customer_Details")
        
        # Add title
        customer_sheet['A1'] = "Customer Details Report"
        customer_sheet['A1'].font = Font(size=16, bold=True)
        customer_sheet.merge_cells('A1:I1')
        customer_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        customer_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        customer_sheet['A2'].font = Font(size=10, italic=True)
        customer_sheet.merge_cells('A2:I2')
        customer_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Add header row at A4
        headers = ["User ID", "First Name", "Last Name", "Username", "Email", "Phone Number", "Address", "Role", "Created At"]
        for col, header in enumerate(headers, start=1):
            cell = customer_sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F1D94B", end_color="F1D94B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        if not customer_df.empty:
            for r_idx, row in enumerate(customer_df.itertuples(), start=5):
                customer_sheet.cell(row=r_idx, column=1, value=row.userID)
                customer_sheet.cell(row=r_idx, column=2, value=row.first_name)
                customer_sheet.cell(row=r_idx, column=3, value=row.last_name)
                customer_sheet.cell(row=r_idx, column=4, value=row.username)
                customer_sheet.cell(row=r_idx, column=5, value=row.email)
                customer_sheet.cell(row=r_idx, column=6, value=row.phone_number)
                customer_sheet.cell(row=r_idx, column=7, value=row.address)
                customer_sheet.cell(row=r_idx, column=8, value=row.role)
                customer_sheet.cell(row=r_idx, column=9, value=str(row.created_at))
        
        # Format sheet
        self._format_sheet(customer_sheet)
    
    def _create_orders_sheet(self, conn, workbook, time_filter, payment_info=None, order_type_info=None):
        """Create orders sheet in the business report by using OrdersOnlyExporter's method"""
        # Create an instance of OrdersOnlyExporter
        orders_exporter = OrdersOnlyExporter()
        
        # Use its _create_orders_sheet method
        orders_exporter._create_orders_sheet(conn, workbook, time_filter, payment_info, order_type_info)
    
    def _create_items_sheet(self, conn, workbook, time_filter):
        """Create items list sheet with detailed order items"""
        # Add time filter to items query
        time_clause = self._get_time_filter_clause(time_filter)
        items_query = f"""
            SELECT 
                o.OrderID,
                o.UserName,
                o.Item_list,
                o.Total_price,
                o.CreatedAT
            FROM `Order` o
            {time_clause}
            ORDER BY o.CreatedAT DESC
        """
        
        # Get data
        cursor = conn.cursor(dictionary=True)
        cursor.execute(items_query)
        orders = cursor.fetchall()
        
        # Create items sheet
        items_sheet = workbook.create_sheet(title="Order_Items")
        
        # Add title
        items_sheet['A1'] = "Order Items Report"
        items_sheet['A1'].font = Font(size=16, bold=True)
        items_sheet.merge_cells('A1:F1')
        items_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        items_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        items_sheet['A2'].font = Font(size=10, italic=True)
        items_sheet.merge_cells('A2:F2')
        items_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Add header row at A4
        headers = ["OrderId", "Customer", "Order Date", "Items List", "Quantity", "Total Price"]
        for col, header in enumerate(headers, start=1):
            cell = items_sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F1D94B", end_color="F1D94B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        if orders:
            row_idx = 5
            for order in orders:
                # Format OrderID to match the image (with leading zeros)
                formatted_order_id = f"#{'0' * (8 - len(str(order['OrderID'])))}{order['OrderID']}"
                
                # Parse the JSON item_list
                try:
                    items_list = json.loads(order['Item_list'])
                except:
                    # For handling malformed JSON or different format
                    items_list = [{"name": "Unknown Item", "quantity": 1}]
                
                # Create a formatted string for the items
                items_str = ""
                quantities_str = ""
                for item in items_list:
                    if isinstance(item, dict) and 'name' in item and 'quantity' in item:
                        items_str += f"{item['name']}\n"
                        quantities_str += f"{item['quantity']}\n"
                
                # Add data to sheet
                items_sheet.cell(row=row_idx, column=1, value=formatted_order_id)
                items_sheet.cell(row=row_idx, column=2, value=order['UserName'])
                items_sheet.cell(row=row_idx, column=3, value=order['CreatedAT'].strftime('%Y-%m-%d %H:%M'))
                
                # Add items with newlines for multiple items
                items_cell = items_sheet.cell(row=row_idx, column=4, value=items_str.strip())
                items_cell.alignment = Alignment(wrapText=True, vertical='top')
                
                # Add quantities with newlines to match items
                qty_cell = items_sheet.cell(row=row_idx, column=5, value=quantities_str.strip())
                qty_cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
                
                # Add total price
                items_sheet.cell(row=row_idx, column=6, value=f"$ {float(order['Total_price']):.2f}")
                
                # Adjust row height based on number of items
                num_items = len(items_list)
                items_sheet.row_dimensions[row_idx].height = max(20, 15 * num_items)
                
                row_idx += 1
        
        # Format sheet
        self._format_sheet(items_sheet)
    
    def _format_sheet(self, sheet):
        """Format an Excel sheet with consistent styling"""
        # Make columns wider based on content
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
        
        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
        
        # Color alternate rows for better readability
        max_row = sheet.max_row
        for row in range(5, max_row + 1):
            if row % 2 == 0:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )


class OrdersOnlyExporter:
    def __init__(self):
        pass
    
    def export_orders(self, time_filter="All Time", payment_info=None, order_type_info=None):
        """Export only orders data to Excel
        
        Args:
            time_filter (str): Time period filter
            payment_info (dict, optional): Dictionary mapping OrderIDs to payment methods from the UI
            order_type_info (dict, optional): Dictionary mapping OrderIDs to order types from the UI
        """
        try:
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Ice'n Spice Orders Export"
            )
            
            if not file_path:
                return
                
            # Connect to database
            conn = mysql.connector.connect(**DB_CONFIG)
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Create orders sheet with detailed information
            self._create_orders_sheet(conn, workbook, time_filter, payment_info, order_type_info)
            
            # Save workbook
            workbook.save(file_path)
            
            # Close database connection
            conn.close()
            
            # Show success message
            messagebox.showinfo("Export Complete", "Orders exported successfully!")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export orders: {str(e)}")
    
    def _get_time_filter_clause(self, time_filter):
        """Generate SQL WHERE clause based on time filter"""
        if time_filter == "This Month":
            return "WHERE YEAR(CreatedAT) = YEAR(CURDATE()) AND MONTH(CreatedAT) = MONTH(CURDATE())"
        elif time_filter == "Last Month":
            return "WHERE YEAR(CreatedAT) = YEAR(DATE_SUB(CURDATE(), INTERVAL 1 MONTH)) AND MONTH(CreatedAT) = MONTH(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))"
        elif time_filter == "This Year":
            return "WHERE YEAR(CreatedAT) = YEAR(CURDATE())"
        else:  # "All Time"
            return ""
    
    def _create_orders_sheet(self, conn, workbook, time_filter, payment_info=None, order_type_info=None):
        """Create orders sheet with all order details"""
        # Add time filter to orders query
        time_clause = self._get_time_filter_clause(time_filter)
        
        # Query to get all order information - without OrderType column
        order_query = f"""
            SELECT 
                OrderID,
                UserName,
                Item_list,
                Total_price,
                Status,
                CreatedAT,
                statusUpdateAt AS ModifiedAT
            FROM `Order`
            {time_clause}
            ORDER BY CreatedAT DESC
        """
        
        # Get data
        cursor = conn.cursor(dictionary=True)
        cursor.execute(order_query)
        orders = cursor.fetchall()
        
        # Create orders sheet
        orders_sheet = workbook.active
        orders_sheet.title = "Orders"
        
        # Add title
        orders_sheet['A1'] = "Orders Report"
        orders_sheet['A1'].font = Font(size=16, bold=True)
        orders_sheet.merge_cells('A1:I1')
        orders_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        orders_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        orders_sheet['A2'].font = Font(size=10, italic=True)
        orders_sheet.merge_cells('A2:I2')
        orders_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Add header row at A4
        headers = [
            "Order ID", 
            "Customer Name", 
            "Order Items", 
            "Total Price", 
            "Payment Method", 
            "Order Status", 
            "Order Type", 
            "Order Date", 
            "Last Modified"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = orders_sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F1D94B", end_color="F1D94B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        if orders:
            row_idx = 5
            for order in orders:
                order_id = order['OrderID']
                
                # Format OrderID with leading zeros
                formatted_order_id = f"#{'0' * (8 - len(str(order_id)))}{order_id}"
                
                # Parse the JSON item_list
                try:
                    items_list = json.loads(order['Item_list'])
                except:
                    items_list = [{"name": "Unknown Item", "quantity": 1}]
                
                # Create a formatted string for items
                items_str = ""
                for item in items_list:
                    if isinstance(item, dict) and 'name' in item and 'quantity' in item:
                        items_str += f"{item['name']} (x{item['quantity']})\n"
                
                # Add data to sheet
                orders_sheet.cell(row=row_idx, column=1, value=formatted_order_id)
                orders_sheet.cell(row=row_idx, column=2, value=order['UserName'])
                
                # Add items with formatting
                items_cell = orders_sheet.cell(row=row_idx, column=3, value=items_str.strip())
                items_cell.alignment = Alignment(wrapText=True, vertical='top')
                
                # Add total with $ sign
                orders_sheet.cell(row=row_idx, column=4, value=f"$ {float(order['Total_price']):.2f}")
                
                # Get payment method from UI data if available
                payment_method = "N/A"
                if payment_info and order_id in payment_info:
                    payment_method = payment_info[order_id]
                orders_sheet.cell(row=row_idx, column=5, value=payment_method)
                
                # Add status with color formatting
                status_cell = orders_sheet.cell(row=row_idx, column=6, value=order['Status'])
                if order['Status'].lower() == 'pending':
                    status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif order['Status'].lower() == 'completed':
                    status_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                status_cell.alignment = Alignment(horizontal='center')
                
                # Get order type from UI data if available
                order_type = "N/A"
                if order_type_info and order_id in order_type_info:
                    order_type = order_type_info[order_id]
                orders_sheet.cell(row=row_idx, column=7, value=order_type)
                
                # Format the dates
                created_at = order['CreatedAT']
                created_str = created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else "Unknown"
                orders_sheet.cell(row=row_idx, column=8, value=created_str)
                
                # Modified date might be None
                modified_at = order['ModifiedAT']
                modified_str = modified_at.strftime('%Y-%m-%d %H:%M:%S') if modified_at else "Not modified"
                orders_sheet.cell(row=row_idx, column=9, value=modified_str)
                
                # Adjust row height based on number of items
                num_items = len(items_list)
                orders_sheet.row_dimensions[row_idx].height = max(20, 15 * num_items)
                
                row_idx += 1
        
        # Format the sheet
        self._format_sheet(orders_sheet)
    
    def _format_sheet(self, sheet):
        """Format an Excel sheet with consistent styling"""
        # Make columns wider based on content
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value).split('\n')[0]))
            adjusted_width = min(max_length + 5, 50)  # Cap at 50 to avoid extremely wide columns
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
        
        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
        
        # Color alternate rows for better readability
        max_row = sheet.max_row
        for row in range(5, max_row + 1):
            if row % 2 == 0:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )


# Example usage:
"""
# In your order screen where you have payment information:
from report_exporter import OrdersOnlyExporter

# Create a dictionary to store payment methods by OrderID
payment_info = {}  # key: OrderID, value: Payment Method
# Fill this dictionary with payment methods from your UI
# For example, when an order is paid:
payment_info[order_id] = "Credit Card"  # or "Cash", "PayPal", etc.

# When exporting orders, pass the payment info to the exporter
orders_exporter = OrdersOnlyExporter()
orders_exporter.export_orders(time_filter="This Month", payment_info=payment_info)
"""